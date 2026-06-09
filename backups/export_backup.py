#!/usr/bin/env python3
"""Exporta as 4 tabelas financeiras para um XLSX antes do reset."""
import os
import sys
import json
import urllib.request
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

URL = "https://zageqyuwodvyxwohpugb.supabase.co"
SRK = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InphZ2VxeXV3b2R2eXh3b2hwdWdiIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NzU1NTk1MiwiZXhwIjoyMDkzMTMxOTUyfQ.PD20LEBm3S2Hnbbjw1wMfJP7eA-ATINSEXu-QRx6hHY"

# Endpoint, ordering, columns ordering
TABLES = [
    ("transacoes", "data_competencia.desc",
     ["data_competencia","tipo","descricao","valor","status","conta_id","categoria_id","entidade_id","forma_pagamento","parcelado","parcela_atual","parcela_total","recorrencia_id","data_pagamento","notas","origem","id","criado_em"]),
    ("recorrencias", "nome.asc",
     ["nome","tipo","valor_padrao","tipo_valor","frequencia","dia_vencimento","dia_semana","conta_id","categoria_id","entidade_id","forma_pagamento","data_inicio","data_fim","ativo","notas","id","criado_em"]),
    ("receitas_brutas", "data_venda.desc",
     ["data_venda","origem","produto_nome","valor_bruto","taxas","valor_liquido","status","data_prevista_pagamento","data_recebimento","metodo_pagamento","parcelas","entidade_id","origem_id","cliente_nome","cliente_email","notas","id","criado_em"]),
    ("greenn_saldos", "capturado_em.desc",
     ["capturado_em","disponivel","pendente","antecipavel","id","created_by"]),
]

def fetch(table, order):
    req = urllib.request.Request(
        f"{URL}/rest/v1/{table}?select=*&order={order}&limit=10000",
        headers={"apikey": SRK, "Authorization": f"Bearer {SRK}"}
    )
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read())

def name_lookups():
    """Pré-carrega lookups pra contas/categorias/entidades/origens facilitarem leitura."""
    out = {}
    for t in ("contas_bancarias","categorias","entidades","origens_receita"):
        req = urllib.request.Request(
            f"{URL}/rest/v1/{t}?select=id,nome&limit=200",
            headers={"apikey": SRK, "Authorization": f"Bearer {SRK}"}
        )
        with urllib.request.urlopen(req) as r:
            out[t] = {row["id"]: row["nome"] for row in json.loads(r.read())}
    return out

def main():
    print("Fetching lookups...")
    look = name_lookups()
    print(f"  contas: {len(look['contas_bancarias'])}, cat: {len(look['categorias'])}, ent: {len(look['entidades'])}, origens: {len(look['origens_receita'])}")

    wb = Workbook()
    wb.remove(wb.active)

    total_rows = 0
    summary = []

    for table, order, cols in TABLES:
        print(f"Fetching {table}...")
        rows = fetch(table, order)
        print(f"  {len(rows)} rows")
        total_rows += len(rows)
        summary.append((table, len(rows)))

        ws = wb.create_sheet(table)
        # Header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="333333")
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(1, ci, col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
        # Data
        for ri, row in enumerate(rows, 2):
            for ci, col in enumerate(cols, 1):
                v = row.get(col)
                # Resolve FKs to readable
                if col == "conta_id" and v in look["contas_bancarias"]:
                    v = look["contas_bancarias"][v]
                elif col == "categoria_id" and v in look["categorias"]:
                    v = look["categorias"][v]
                elif col == "entidade_id" and v in look["entidades"]:
                    v = look["entidades"][v]
                elif col == "origem_id" and v in look["origens_receita"]:
                    v = look["origens_receita"][v]
                # Format
                if isinstance(v, (dict, list)):
                    v = json.dumps(v, ensure_ascii=False)
                ws.cell(ri, ci, v)
        # Auto-width
        for ci, col in enumerate(cols, 1):
            max_len = max(
                [len(str(col))] + [len(str(ws.cell(ri+2, ci).value or "")) for ri in range(min(50, len(rows)))]
            )
            ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len + 2, 12), 40)
        ws.freeze_panes = "A2"

    # Summary sheet first
    ws_sum = wb.create_sheet("_resumo", 0)
    ws_sum["A1"] = "Backup financeiro v2"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws_sum["A4"] = "Tabela"
    ws_sum["B4"] = "Linhas"
    ws_sum["A4"].font = Font(bold=True)
    ws_sum["B4"].font = Font(bold=True)
    for i, (t, n) in enumerate(summary, 5):
        ws_sum.cell(i, 1, t)
        ws_sum.cell(i, 2, n)
    ws_sum.cell(5 + len(summary), 1, "TOTAL").font = Font(bold=True)
    ws_sum.cell(5 + len(summary), 2, total_rows).font = Font(bold=True)
    ws_sum.column_dimensions["A"].width = 25
    ws_sum.column_dimensions["B"].width = 12

    out_path = sys.argv[1] if len(sys.argv) > 1 else "financeiro_backup.xlsx"
    wb.save(out_path)
    print(f"\nSaved: {out_path}")
    print(f"Total rows backed up: {total_rows}")

if __name__ == "__main__":
    main()
